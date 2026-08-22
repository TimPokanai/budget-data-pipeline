"""Builds the synthetic three-sheet workbook CI ingests in place of a real
monthly workbook (never committed -- see .gitignore and PROJECT_PLAN.md's
source-data rationale).

Generated fresh by every CI run rather than committed as a binary fixture:
one less binary diff to review, and the dataset in fixture_data.py is the
single source of truth a reviewer actually reads.

Usage:
    python scripts/ci/generate_fixture_workbook.py [output_path]
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook

from fixture_data import BUDGETS, CATEGORY_NAMES, FILENAME, TRANSACTIONS, Txn


def build_workbook(
    path: Path,
    transactions: list[Txn] = TRANSACTIONS,
    budgets: list[tuple[str, float]] = BUDGETS,
    category_names: list[str] = CATEGORY_NAMES,
) -> Path:
    """Writes a workbook matching the real one's sheet/column structure
    (see docs/phase-1-schema-design.md's source-structure table) to `path`.
    """
    wb = Workbook()

    ws_tracker = wb.active
    ws_tracker.title = "Expense Tracker"
    ws_tracker.append(["Date", "Description", "Category", "Amount (CAD)"])
    for t in transactions:
        # `None` values (the blank-description corrupted case) are passed
        # straight through -- ws.append() writes a genuinely blank cell for
        # None, not the string "None". See README's openpyxl gotcha note:
        # this only bites when using cell(value=None) instead of a direct
        # assignment/append, which is what happens here.
        ws_tracker.append([t.txn_date, t.description, t.category, t.amount])

    ws_summary = wb.create_sheet("Monthly Budget Summary")
    ws_summary.append(["Category", "Planned (CAD)", "Actual (CAD)", "Difference (CAD)"])
    first_row = 2
    for i, (category, planned) in enumerate(budgets):
        r = first_row + i
        # Formula text, never evaluated (openpyxl has no calc engine, and
        # nothing downstream reads these two columns -- see
        # docs/phase-2-ingestion.md, "Actual/Difference become a dbt
        # model"). Written anyway so the fixture is structurally honest
        # about what the real workbook contains.
        actual_formula = f"=SUMIFS('Expense Tracker'!$D:$D,'Expense Tracker'!$C:$C,$A{r})"
        diff_formula = f"=C{r}-B{r}"
        ws_summary.append([category, planned, actual_formula, diff_formula])
    total_row = first_row + len(budgets)
    ws_summary.append([
        "Total",
        f"=SUM(B{first_row}:B{total_row - 1})",
        f"=SUM(C{first_row}:C{total_row - 1})",
        f"=SUM(D{first_row}:D{total_row - 1})",
    ])

    ws_categories = wb.create_sheet("Categories")
    ws_categories.append(["Names"])
    for name in category_names:
        ws_categories.append([name])

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def default_output_path() -> Path:
    # scripts/ci/ -> repo root -> tests/fixtures/
    repo_root = Path(__file__).resolve().parents[2]
    return repo_root / "tests" / "fixtures" / FILENAME


if __name__ == "__main__":
    out_path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_output_path()
    written = build_workbook(out_path)
    print(f"Wrote fixture workbook: {written}")
